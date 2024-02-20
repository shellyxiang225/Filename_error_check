import os
import re
import tkinter as tk
import CheckError as ttk
from tkinter import filedialog
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border


# 跳出选框，引入文件路径
root = tk.Tk()
root.withdraw()
FolderPath = filedialog.askdirectory()
dir = FolderPath
# root指根目录，dir指子目录，file文件名
list = []


for root, dir, file in os.walk(dir):
    # 在file列表中遍历得到文件名
    for d in file:
        file_path = os.path.join(root, file)
        # pattern解释：文件名由中文名+英文+数字组成文件名，（数字和.)组成K3代码，所有字符为英文符号，扩展名暂无规范
        if re.search(r"^[\u4e00-\u9fa5a-zA-z0-9]+\(\d{1}\.\d{2}\.\d{4}\-\d{3}\)\.\w+$", d):
            print(d, 'T')
            list.append(d)
            list.append('T')
            list.append(file_path)

        else:
            print(d, 'F')
            list.append(d)
            list.append('F')
            list.append(file_path)
print(str(list))

wb = openpyxl.Workbook()
ws = wb.create_sheet('sheet', 0)
font = Font(name='宋体', size=12, color='000000', bold=True, underline='single')
fill = PatternFill(start_color='c8edfc', end_color='c8edfc', fill_type='solid')
ws.cell(row=1, column=1).value = '文件名称'
ws.cell(row=1, column=2).value = '正误'
ws['A1'].font = font
ws['A1'].fill = fill
ws['B1'].font = font
ws['B1'].fill = fill
wb.save('check_new.xlsx')
wb.close()
workbook = load_workbook(filename="check_new.xlsx")
sheet = workbook.active
# 列表里每2个数据存入excel
for i in range(0, len(list), 2):
    row = [list[i], list[i + 1]]
    sheet.append(row)
    workbook.save('check_new.xlsx')
    wb.close()

input('Press <enter>')
